#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, Style, Alignment
from openpyxl.styles.colors import BLUE
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation, ValidationType

from datetime import datetime

# My modules
import config as cf

def dataval_any(ws, range):
    # Create a data-validation object with list validation
    dv = DataValidation(type='date')

    # Add the data-validation object to the worksheet
    ws.add_data_validation(dv)

    # Or, apply the validation to a range of cells
    dv.ranges.append(range)

def dataval_list(ws, enum, range):
    # Create a data-validation object with list validation
    dvl = DataValidation(type="list", formula1=enum, allow_blank=True)

    # Optionally set a custom error message
    dvl.error ='Your entry is not in the list'
    dvl.errorTitle = 'Invalid Entry'

    # Optionally set a custom prompt message
    dvl.prompt = 'Please select from the list'
    dvl.promptTitle = 'List Selection'

    # Add the data-validation object to the worksheet
    ws.add_data_validation(dvl)
    
    # Or, apply the validation to a range of cells
    dvl.ranges.append(range)


xlfile = r'/Users/sroberts/Box Sync/Python/Sandbox/FDRP2.xlsx'

# Open the spreadsheet
try:
    wb = load_workbook(xlfile)
    print('Opened existing file :', xlfile)
    existing_doc = True
except:
    wb = openpyxl.Workbook()
    print('Created new file :', xlfile)
    existing_doc = False
    
print(wb.get_sheet_names())

# select the worksheet
ws = wb.get_sheet_by_name('Discussion Items')

# Entry Type
et_enum = '=Enums!$A$3:$A$5'
et_range = 'L5:L1000'
dataval_list(ws, et_enum, et_range)

# Discussion Priority
dp_enum = '=Enums!$B$3:$B$5'
dp_range = 'M5:M1000'
dataval_list(ws, dp_enum, dp_range)

# Review Significance
rs_enum = '=Enums!$C$3:$C$5'
rs_range = 'N5:N1000'
dataval_list(ws, rs_enum, rs_range)

# Review Charge
rc_enum = "=Enums!$D$3:$D$16"
rc_range = "O5:O1000"
dataval_list(ws, rc_enum, rc_range)

# Related Document
rd_enum = '=Enums!$E$3:$E$48'
rd_range = 'P5:P1000'
dataval_list(ws, rd_enum, rd_range)

# Write the sheet out.  If you now open the sheet in Excel, you'll find that
# the cells have data-validation applied.
wb.save(xlfile)